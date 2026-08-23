from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import MAX_PARTICIPANTS

HEADERS = ["Fecha de inscripción", "Total integrantes"]
for i in range(1, MAX_PARTICIPANTS + 1):
    HEADERS += [
        f"P{i} - Nombre" + (" (líder)" if i == 1 else ""),
        f"P{i} - Documento",
        f"P{i} - Institución",
        f"P{i} - Correo institucional",
        f"P{i} - Celular",
    ]


def build_registrations_workbook(registrations) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inscripciones 35mm"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for registration in registrations:
        participants_by_position = {p.position: p for p in registration.participants.all()}
        row = [
            registration.created_at.strftime("%Y-%m-%d %H:%M"),
            len(participants_by_position),
        ]
        for i in range(1, MAX_PARTICIPANTS + 1):
            p = participants_by_position.get(i)
            if p:
                row += [p.full_name, p.document_id, p.institution, p.institutional_email, p.phone]
            else:
                row += ["", "", "", "", ""]
        ws.append(row)

    for i, _ in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    return wb


def registrations_filename() -> str:
    return f"inscripciones_35mm_{datetime.now():%Y%m%d_%H%M}.xlsx"
